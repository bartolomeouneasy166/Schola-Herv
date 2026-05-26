"""
Metadata builder – enriches papers with journal metrics and builds Excel reports.
"""

import asyncio
import csv
import re
from pathlib import Path
from typing import List, Optional, Dict

import aiohttp
import openpyxl
from openpyxl.styles import Font, Alignment

# Default columns for the output Excel file
DEFAULT_COLUMNS = [
    "Number",
    "Title",
    "Author",
    "Year",
    "Journal",
    "Category (Q1-Q4)",
    "Abstract",
    "DOI",
    "Impact Factor",
]


def load_scimago_csv(csv_path: Optional[str]) -> Optional[Dict[str, Dict[str, str]]]:
    """
    Load SCImago Journal Rank CSV and return a dict mapping ISSN -> {title, impact_factor, quartile}.
    The CSV is expected to have columns: 'Issn', 'Title', 'Cites / Doc. (2years)', 'SJR Quartile'.
    If the file is missing or unreadable, returns None.
    """
    if not csv_path or not Path(csv_path).exists():
        return None

    issn_map = {}
    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                issn_str = row.get("Issn", "").strip()
                title = row.get("Title", "").strip()
                impact = row.get("Cites / Doc. (2years)", "").strip()
                quartile = row.get("SJR Quartile", "").strip()
                if not issn_str or not title:
                    continue
                # Some rows contain multiple ISSNs separated by comma or space
                issns = re.split(r"[,\s]+", issn_str)
                for issn in issns:
                    issn = issn.strip()
                    if issn:
                        issn_map[issn.lower()] = {
                            "title": title,
                            "impact_factor": impact,
                            "quartile": quartile,
                        }
        return issn_map
    except Exception:
        return None


def enrich_with_journal_metrics(papers: List[dict],
                                scimago_data: Optional[Dict[str, Dict[str, str]]]) -> List[dict]:
    """
    Add 'quartile' and 'impact_factor' fields to each paper dict using the journal name
    or ISSN (if available). Falls back to empty string if no match.
    """
    if not scimago_data:
        for p in papers:
            p["quartile"] = ""
            p["impact_factor"] = ""
        return papers

    for paper in papers:
        quartile = ""
        impact = ""
        journal_name = (paper.get("journal") or "").strip().lower()
        issn = (paper.get("issn") or "").strip().lower()

        if issn and issn in scimago_data:
            quartile = scimago_data[issn]["quartile"]
            impact = scimago_data[issn]["impact_factor"]
        elif journal_name:
            # Simple case‑insensitive name match
            for issn_key, info in scimago_data.items():
                if info["title"].lower() == journal_name:
                    quartile = info["quartile"]
                    impact = info["impact_factor"]
                    break
        paper["quartile"] = quartile or ""
        paper["impact_factor"] = impact or ""
    return papers


async def resolve_missing_dois(papers: List[dict]) -> List[dict]:
    """
    For papers without a DOI, query Crossref using title + first author
    to find the DOI. Modifies papers in place and returns them.
    """
    if not papers:
        return papers

    async def resolve_one(session, paper):
        if paper.get("doi"):
            return
        title = paper.get("title", "").strip()
        if not title:
            return

        # Extract first author's last name for more accurate search
        authors = paper.get("authors", [])
        first_author = ""
        if isinstance(authors, list) and authors:
            first_author = authors[0].split()[-1] if authors[0] else ""
        elif isinstance(authors, str) and authors:
            first_author = authors.split(";")[0].strip().split()[-1] if ";" in authors else authors.split()[-1]

        query = title
        if first_author:
            query += f" {first_author}"

        url = "https://api.crossref.org/works"
        params = {
            "query.bibliographic": query,
            "rows": 1,
            "sort": "relevance",
        }
        try:
            async with session.get(url, params=params, timeout=10) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    items = data.get("message", {}).get("items", [])
                    if items:
                        best = items[0]
                        doi = best.get("DOI")
                        if doi:
                            paper["doi"] = doi
                            # Optionally fill other missing fields
                            if not paper.get("journal"):
                                container = best.get("container-title", [])
                                paper["journal"] = container[0] if container else ""
                            if not paper.get("year"):
                                year_parts = best.get("published-print", {}).get("date-parts", [[None]])[0]
                                if year_parts[0]:
                                    paper["year"] = year_parts[0]
                            if not paper.get("abstract"):
                                paper["abstract"] = best.get("abstract", "")
        except Exception:
            pass

    async with aiohttp.ClientSession(
        headers={"User-Agent": "Schola-Herv/1.0 (mailto:researcher@example.com)"}
    ) as session:
        tasks = [resolve_one(session, p) for p in papers]
        await asyncio.gather(*tasks)
    return papers


def build_excel(papers: List[dict],
                output_path: Path,
                include_pdf_col: bool = False) -> None:
    """
    Write an Excel file with the standard columns.
    If include_pdf_col is True, an extra 'PDF Downloaded' column is added.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schola-Herv Survey"

    columns = DEFAULT_COLUMNS.copy()
    if include_pdf_col:
        columns.append("PDF Downloaded")

    # Header formatting
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, paper in enumerate(papers, 2):
        # Number
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        # Title
        ws.cell(row=row_idx, column=2, value=paper.get("title", ""))
        # Author – join list with semicolons
        authors = paper.get("authors", [])
        if isinstance(authors, list):
            authors_str = "; ".join(authors)
        else:
            authors_str = str(authors)
        ws.cell(row=row_idx, column=3, value=authors_str)
        # Year
        ws.cell(row=row_idx, column=4, value=paper.get("year"))
        # Journal
        ws.cell(row=row_idx, column=5, value=paper.get("journal") or "")
        # Category (quartile)
        ws.cell(row=row_idx, column=6, value=paper.get("quartile", ""))
        # Abstract
        ws.cell(row=row_idx, column=7, value=paper.get("abstract") or "")
        # DOI
        ws.cell(row=row_idx, column=8, value=paper.get("doi") or "")
        # Impact Factor
        ws.cell(row=row_idx, column=9, value=paper.get("impact_factor", ""))

        if include_pdf_col:
            ws.cell(row=row_idx, column=10,
                    value=paper.get("pdf_downloaded", False))

    # Auto-adjust column widths (simple heuristic)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_path)